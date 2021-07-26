#! /usr/bin/env python 3.7 (3394)
#coding=utf-8
# Compiled at: 1969-12-31 18:00:00
#Powered by BugScaner
#http://tools.bugscaner.com/
#如果觉得不错,请分享给你朋友使用吧!
u"""
Created on 2019\u5e7411\u67089\u65e5

@author: \u778c\u7761\u87f2\u5b50
"""
import openpyxl
from openpyxl import utils, Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

class Excel07:

    def __init__(self):
        self.excel = None
        self.path = None

    def CreateExcel(self, sPath):
        self.excel = Workbook()
        self.excel['Sheet'].title = 'Sheet1'
        self.excel.save(sPath)
        self.path = sPath
        return self

    def OpenExcel(self, sPath):
        stuff = str.lower(sPath[-4:])
        if stuff == 'xlsm':
            self.excel = openpyxl.load_workbook(sPath, keep_vba=True)
        else:
            self.excel = openpyxl.load_workbook(sPath)
        self.path = sPath
        return self

    def Save(self, bSave=True):
        if bSave:
            self.excel.save(self.path)

    def CreateSheet(self, strSheetName, strWhere='after', bSave=False):
        asheet = self.excel.active
        asheet = self.excel.index(asheet)
        if strWhere.lower() == 'after':
            strWhere = asheet + 1
        elif strWhere.lower() == 'before':
                strWhere = asheet
        else:
            self._getException(strWhere)
        if strWhere:
            self.excel.create_sheet(strSheetName, strWhere)
        else:
            self.excel.create_sheet(strSheetName)
        self.Save(bSave)

    def GetSheetsName(self):
        return self.excel.sheetnames

    def SheetRename(self, sheet, strNewName, bSave=False):
        self._getSheet(sheet).title = strNewName
        self.Save(bSave)

    def CopySheet(self, sheet, strNewSheetName, bSave=False):
        tempSheet = self.excel.copy_worksheet(self._getSheet(sheet))
        tempSheet.title = strNewSheetName
        self.Save(bSave)

    def DeleteSheet(self, sheet, bSave=False):
        self.excel.remove(self._getSheet(sheet))
        self.Save(bSave)

    def ActiveSheet(self, sheet):
        Workbook.active = self._getSheet(sheet)

    def WriteCell(self, sheet, strCell, data, bSave=False):
        cell = self._getCellOrRange(sheet, strCell)
        cell.value = data
        self.Save(bSave)

    def ReadCell(self, sheet, strCell):
        cell = self._getCellOrRange(sheet, strCell)
        return cell.value

    def WriteRow(self, sheet, strCell, data, bSave=False):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strCell)
        column, row = utils.cell.coordinate_from_string(tempCells)
        column = utils.column_index_from_string(column)
        for k in range(len(data)):
            cell = '%s%d' % (utils.get_column_letter(column + k), row)
            tempSheet[cell].value = data[k]

        self.Save(bSave)

    def WriteColumn(self, sheet, strCell, data, bSave=False):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strCell)
        column, row = utils.cell.coordinate_from_string(tempCells)
        for k in range(len(data)):
            cell = '%s%d' % (column, row + k)
            tempSheet[cell].value = data[k]

        self.Save(bSave)

    def ReadRow(self, sheet, strCell):
        strRange = self._getRow(sheet, strCell)
        return self.ReadRange(sheet, strRange)[0]

    def ReadColumn(self, sheet, strCell):
        strRange = self._getColumn(sheet, strCell)
        tempRange = self.ReadRange(sheet, strRange)
        return [y for y in tempRange]

    def InsertRow(self, sheet, strCell, data, bSave=False):
        tempSheet = self._getSheet(sheet)
        if type(strCell) == int:
            tempSheet.insert_rows(strCell)
            strCell = '%s%d' % ('A', strCell)
        else:
            tempCells = self._getCells(strCell)
            column, row = utils.cell.coordinate_from_string(tempCells)
            tempSheet.insert_rows(row)
            strCell = tempCells
        self.WriteRow(sheet, strCell, data, bSave)

    def InsertColumn(self, sheet, strCell, data, bSave=False):
        tempSheet = self._getSheet(sheet)
        if type(strCell) == int:
            tempSheet.insert_cols(strCell)
            strCell = '%s%d' % (utils.get_column_letter(strCell), 1)
        else:
            tempCells = self._getCells(strCell)
            column, row = utils.cell.coordinate_from_string(tempCells)
            tempSheet.insert_cols(utils.column_index_from_string(column))
            strCell = tempCells
        self.WriteColumn(sheet, strCell, data, bSave)

    def MergeRange(self, sheet, strRange, option=True, bSave=False):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strRange)
        if option:
            tempSheet.merge_cells(tempCells)
        else:
            tempSheet.unmerge_cells(tempCells)
        self.Save(bSave)

    def CloseExcel(self, bSave=False):
        self.Save(bSave)
        self.excel.close()

    def ReadRange(self, sheet, strRange):
        cells = self._getCellOrRange(sheet, strRange)
        res = []
        for r in cells:
            res.append([v.value for v in r])

        return res

    def GetRowsCount(self, sheet):
        return self._getSheet(sheet).max_row

    def GetColumnsCount(self, sheet):
        return self._getSheet(sheet).max_column

    def DeleteRow(self, sheet, strCell, bSave=False):
        tempSheet = self._getSheet(sheet)
        if type(strCell) == int:
            tempSheet.delete_rows(strCell)
        else:
            tempCells = self._getCells(strCell)
            column, row = utils.cell.coordinate_from_string(tempCells)
            tempSheet.delete_rows(row)
        self.Save(bSave)

    def DeleteColumn(self, sheet, strCell, bSave=False):
        tempSheet = self._getSheet(sheet)
        if type(strCell) == int:
            tempSheet.delete_cols(strCell)
        else:
            tempCells = self._getCells(strCell)
            column, row = utils.cell.coordinate_from_string(tempCells)
            tempSheet.delete_cols(utils.column_index_from_string(column))
        self.Save(bSave)

    def InsertImage(self, sheet, strCell, sFilePath, fWidth, fHeight, bSave=False):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strCell)
        column, row = utils.cell.coordinate_from_string(tempCells)
        img = Image(sFilePath)
        img.width, img.height = fWidth, fHeight
        tempSheet.column_dimensions[column].width = fWidth
        tempSheet.row_dimensions[row].height = fHeight
        tempSheet.add_image(img, tempCells)
        self.Save(bSave)

    def DeleteImage(self, sheet, objPic, bSave=False):
        tempSheet = self._getSheet(sheet)
        del tempSheet._images[objPic]
        self.Save(bSave)

    def WriteRange(self, sheet, strCell, data, bSave=False):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strCell)
        column, row = utils.cell.coordinate_from_string(tempCells)
        column = utils.column_index_from_string(column)
        for rg in range(len(data)):
            for r in range(len(data[rg])):
                cell = '%s%d' % (utils.get_column_letter(column + r), row + rg)
                tempSheet[cell].value = data[rg][r]

        self.Save(bSave)

    def ClearRange(self, sheet, strRange, bClearFormat=True, bSave=False):
        cells = self._getCellOrRange(sheet, strRange)
        for row in cells:
            for cell in row:
                cell.value = None
                if bClearFormat:
                    cell.font = Font()
                    cell.fill = PatternFill()

        self.Save(bSave)

    def DeleteRange(self, sheet, strRange, bSave=False):
        tempRange = self._getCells(strRange)
        min_col, min_row, max_col, max_row1 = utils.cell.range_boundaries(tempRange)
        max_row = self.GetRowsCount(sheet)
        tempRange = self._getCells([[max_row1, min_col], [max_row, max_col]])
        tempSheet = self._getSheet(sheet)
        tempSheet.move_range(tempRange, rows=min_row - max_row1, cols=0, translate=True)
        self.ClearRange(sheet, [[max_row - (max_row1 - min_row), min_col], [max_row, max_col]], True, False)
        self.Save(bSave)

    def SetRowHeight(self, sheet, strCell, fHeight, bSave=False):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strCell)
        column, row = utils.cell.coordinate_from_string(tempCells)
        tempSheet.row_dimensions[row].height = fHeight
        self.Save(bSave)

    def SetColumnWidth(self, sheet, strCell, fWidth, bSave=False):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strCell)
        column, row = utils.cell.coordinate_from_string(tempCells)
        tempSheet.column_dimensions[column].width = fWidth
        self.Save(bSave)

    def SetCellFontColor(self, sheet, strCell, listColor, bSave=False):
        cell = self._getCellOrRange(sheet, strCell)
        cell.font = Font(color=self._getColor(listColor))
        self.Save(bSave)

    def SetRangeFontColor(self, sheet, strRange, listColor, bSave=False):
        cells = self._getCellOrRange(sheet, strRange)
        tempColor = Font(color=self._getColor(listColor))
        for row in cells:
            for cell in row:
                cell.font = tempColor

        self.Save(bSave)

    def SetCellColor(self, sheet, strCell, listColor, bSave=False):
        cell = self._getCellOrRange(sheet, strCell)
        cell.fill = PatternFill('solid', fgColor=self._getColor(listColor))
        self.Save(bSave)

    def SetRangeColor(self, sheet, strRange, listColor, bSave=False):
        cells = self._getCellOrRange(sheet, strRange)
        tempColor = PatternFill('solid', fgColor=self._getColor(listColor))
        for row in cells:
            for cell in row:
                cell.fill = tempColor

        self.Save(bSave)

    def _getSheet(self, sheet):
        if type(sheet) == str:
            return self.excel[sheet]
        if type(sheet) == int:
            return self.excel.worksheets[sheet]

    def _getCells(self, strCell):
        if type(strCell) == str:
            return strCell.upper()
        if type(strCell) == list:
            if len(strCell) == 2:
                valueTemp = strCell[0]
                expTemp = ''
                if type(valueTemp) == int:
                    expTemp = '%s%d' % (utils.get_column_letter(strCell[1]), valueTemp)
                else:
                    if type(valueTemp) == list:
                        expTemp = ['%s%d' % (utils.get_column_letter(r[1]), r[0]) for r in strCell]
                        expTemp = ':'.join(expTemp)
                    else:
                        self._getException(strCell)
                return expTemp.upper()
            self._getException(strCell)

    def _getColumn(self, sheet, strCell):
        tempSheet = self._getSheet(sheet)
        if type(strCell) == str:
            column, row = utils.cell.coordinate_from_string(strCell)
            return '%s:%s%d' % (strCell, column, tempSheet.max_row)
        if type(strCell) == list:
            tempList = list(filter(lambda x: type(x) == int, strCell))
            if len(tempList) == 2:
                strRange = []
                strRange.append(strCell)
                strRange.append([tempSheet.max_row, strCell[1]])
                return strRange
            self._getException(strCell)

    def _getRow(self, sheet, strCell):
        tempSheet = self._getSheet(sheet)
        if type(strCell) == str:
            column, row = utils.cell.coordinate_from_string(strCell)
            return '%s:%s%d' % (strCell, utils.get_column_letter(tempSheet.max_column), row)
        if type(strCell) == list:
            tempList = list(filter(lambda x: type(x) == int, strCell))
            if len(tempList) == 2:
                strRange = []
                strRange.append(strCell)
                strRange.append([strCell[0], tempSheet.max_column])
                return strRange
            self._getException(strCell)

    def _getCellOrRange(self, sheet, strCell):
        tempSheet = self._getSheet(sheet)
        tempCells = self._getCells(strCell)
        return tempSheet[tempCells]

    def _getColor(self, color):
        if type(color) == str:
            if len(color) == 6:
                return color.upper()
        if type(color) == list:
            if len(color) == 3:
                tempList = list(filter(lambda x: type(x) == int and 0 <= x <= 255, color))
                if len(tempList) == 3:
                    return ''.join([str(hex(r))[-2:].replace('x', '0').upper() for r in color])
                self._getException(color)
            self._getException(color)

    def _getException(self, msg):
        raise Exception(u'\u53c2\u6570\u9519\u8bef:%s' % str(msg))


# if __name__ == '__main__':
    # xls = Excel07()
    # xls.OpenExcel(u'C:\\Users\\Administrator\\Desktop\\\u63d2\u4ef67.xlsx')
    # xls.DeleteRange(0, 'b5:d8', True)
    # xls.SetRangeColor('Sheet1', 'B24:e27', [255, 255, 0], True)
    # xls.CloseExcel()