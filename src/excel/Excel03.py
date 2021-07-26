__doc__ = '\nCreated on 2019年11月9日\n\n@author: 瞌睡蟲子\n'
import xlwt, xlrd
from xlutils.compat import xrange
from xlutils.filter import XLWTWriter, XLRDReader, BaseFilter, BaseWriter, process
from openpyxl import utils

class Excel03:

    def __init__(self):
        self._Excel03__wt = None
        self._Excel03__rd = None
        self._Excel03__style_list = None
        self._Excel03__path = None

    def CreateExcel(self, sPath):
        self._Excel03__wt = xlwt.Workbook()
        self._Excel03__wt.add_sheet('Sheet1', cell_overwrite_ok=True)
        self._Excel03__wt.save(sPath)
        return self.OpenExcel(sPath)

    def OpenExcel(self, sPath):
        self._Excel03__rd = xlrd.open_workbook(sPath, formatting_info=True)
        w = XLWTWriter()
        process(XLRDReader(self._Excel03__rd, 'unknown.xls'), w)
        self._Excel03__wt = w.output[0][1]
        self._Excel03__style_list = w.style_list
        self._Excel03__path = sPath
        return self

    def Save(self, bSave=True):
        if bSave:
            self._Excel03__wt.save(self._Excel03__path)

    def CloseExcel(self, bSave=False):
        self.Save(bSave)
        self._Excel03__rd.release_resources()

    def CreateSheet(self, strSheetName, strWhere='after', bSave=False):
        asheet = self._Excel03__wt.get_active_sheet()
        if strWhere.lower() == 'after':
            strWhere = asheet + 1
        elif strWhere.lower() == 'before':
            strWhere = asheet
        else:
            self._getException(strWhere)
        self._Excel03__wt.add_sheet(strSheetName, cell_overwrite_ok=True)
        self.Save(bSave)
        if strWhere:
            if type(strWhere) == int:
                self._Excel03__handleSheet(strSheetName, mvIndex=strWhere, option=1)
                self.Save(bSave)

    def GetSheetsName(self):
        return self._Excel03__rd.sheet_names()

    def SheetRename(self, sheet, strNewName, bSave=False):
        self._getWtSheet(sheet).name = strNewName
        self.Save(bSave)

    def CopySheet(self, sheet, strNewSheetName, bSave=False):
        self._Excel03__handleSheet(sheet, wtSheetname=strNewSheetName, option=2)
        self.Save(bSave)

    def DeleteSheet(self, sheet, bSave=False):
        self._Excel03__handleSheet(sheet, option=3)
        self.Save(bSave)

    def ActiveSheet(self, sheet):
        tempSheet = self._getWtSheet(sheet)
        index = self._Excel03__wt.sheet_index(tempSheet.name)
        self._Excel03__wt.set_active_sheet(index)

    def WriteCell(self, sheet, strCell, data, bSave=False):
        row, col, value, style = self._copyStyle(sheet, strCell)
        tempSheet = self._getWtSheet(sheet)
        tempSheet.write(row, col, data, style)
        self.Save(bSave)

    def ReadCell(self, sheet, strCell):
        tempSheet = self._getRdSheet(sheet)
        col, row = self._getCells(strCell)
        return tempSheet.cell_value(row, col)

    def WriteRow(self, sheet, strCell, data, bSave=False):
        tempSheet = None
        col, row = self._getCells(strCell)
        isrefresh = True
        for r in range(len(data)):
            row1, col1, value, style = self._copyStyle(sheet, [row + 1, col + r + 1], isrefresh)
            if isrefresh:
                tempSheet = self._getWtSheet(sheet)
                isrefresh = False
            tempSheet.write(row1, col1, data[r], style)

        self.Save(bSave)

    def WriteColumn(self, sheet, strCell, data, bSave=False):
        tempSheet = None
        col, row = self._getCells(strCell)
        isrefresh = True
        for r in range(len(data)):
            row1, col1, value, style = self._copyStyle(sheet, [row + r + 1, col + 1], isrefresh)
            if isrefresh:
                tempSheet = self._getWtSheet(sheet)
                isrefresh = False
            tempSheet.write(row1, col1, data[r], style)

        self.Save(bSave)

    def ReadRow(self, sheet, strCell):
        col, row = self._getCells(strCell)
        return self._getRdSheet(sheet).row_values(row, col)

    def ReadColumn(self, sheet, strCell):
        col, row = self._getCells(strCell)
        return self._getRdSheet(sheet).col_values(col, row)

    def InsertRow(self, sheet, strCell, data, bSave=False):
        max_col = self.GetColumsCount(sheet)
        min_col, min_row = self._getCells(strCell)
        self._Excel03__handleCell(sheet, [min_row, 0, min_row + 1, max_col], way=False)
        self.WriteRow(sheet, strCell, data, False)
        self.Save(bSave)

    def InsertColumn(self, sheet, strCell, data, bSave=False):
        max_row = self.GetRowsCount(sheet)
        min_col, min_row = self._getCells(strCell)
        self._Excel03__handleCell(sheet, [0, min_col, max_row, min_col + 1])
        self.WriteColumn(sheet, strCell, data, bSave)
        self.Save(bSave)

    def MergeRange(self, sheet, strRange, option=True, bSave=False):
        tempSheet = self._getWtSheet(sheet)
        min_col, min_row, max_col, max_row = self._getCells(strRange)
        if option:
            tempSheet.merge(min_row, max_row - 1, min_col, max_col - 1)
        else:
            self._Excel03__handleMerge(sheet, [min_col, min_row, max_col, max_row])
        self.Save(bSave)

    def ReadRange(self, sheet, strRange):
        tempSheet = self._getRdSheet(sheet)
        min_col, min_row, max_col, max_row = self._getCells(strRange)
        res = []
        for row in range(min_row, max_row):
            res.append(tempSheet.row_values(row, min_col, max_col))

        return res

    def GetRowsCount(self, sheet):
        return self._getRdSheet(sheet).nrows

    def GetColumnsCount(self, sheet):
        return self._getRdSheet(sheet).ncols

    def DeleteRow(self, sheet, strCell, bSave=False):
        max_col = self.GetColumsCount(sheet)
        min_col, min_row = self._getCells(strCell)
        self._Excel03__handleCell(sheet, [min_row, 0, min_row + 1, max_col], way=False, option=False)
        self.Save(bSave)

    def DeleteColumn(self, sheet, strCell, bSave=False):
        max_row = self.GetRowsCount(sheet)
        min_col, min_row = self._getCells(strCell)
        self._Excel03__handleCell(sheet, [0, min_col, max_row, min_col + 1], option=False)
        self.Save(bSave)

    def InsertImage(self, sheet, strCell, sFilePath, fWidth, fHeight, bSave=False):
        col, row = self._getCells(strCell)
        tempSheet = self._getWtSheet(sheet)
        tempSheet.insert_bitmap(sFilePath, row, col, 0, 0, fWidth, fHeight)
        self.Save(bSave)

    def DeleteImage(self, sheet, objPic, bSave=False):
        raise Exception('xls暂未实现图片删除！')

    def WriteRange(self, sheet, strCell, data, bSave=False):
        tempSheet = None
        col, row = self._getCells(strCell)
        isrefresh = True
        for rg in range(len(data)):
            for r in range(len(data[rg])):
                row1, col1, value, style = self._copyStyle(sheet, [row + rg + 1, col + r + 1], isrefresh)
                if not tempSheet:
                    tempSheet = self._getWtSheet(sheet)
                isrefresh = False
                tempSheet.write(row1, col1, data[rg][r], style)

        self.Save(bSave)

    def ClearRange(self, sheet, strRange, bClearFormat=True, bSave=False):
        min_col, min_row, max_col, max_row = self._getCells(strRange)
        isrefresh = True
        tempSheet = self._getWtSheet(sheet)
        for row in range(min_row, max_row):
            for col in range(min_col, max_col):
                if bClearFormat:
                    tempSheet.write(row, col)
                else:
                    row, col, value, style = self._copyStyle(sheet, [row, col], isrefresh)
                    if isrefresh:
                        tempSheet = self._getWtSheet(sheet)
                        isrefresh = False
                    tempSheet.write(row, col, '', style)

        self.Save(bSave)

    def DeleteRange(self, sheet, strRange, bSave=False):
        min_col, min_row, max_col, max_row = self._getCells(strRange)
        self._Excel03__handleCell(sheet, [min_row, min_col, max_row, max_col], way=False, option=False)
        self.Save(bSave)

    def SetRowHeight(self, sheet, strCell, fHeight, bSave=False):
        tempSheet = self._getWtSheet(sheet)
        col, row = self._getCells(strCell)
        hestyle = xlwt.easyxf('font:height ' + str(72 * fHeight))
        tempSheet.row(row).set_style(hestyle)
        self.Save(bSave)

    def SetColumnWidth(self, sheet, strCell, fWidth, bSave=False):
        tempSheet = self._getWtSheet(sheet)
        col, row = self._getCells(strCell)
        tempSheet.col(col).width = int(219.42857142857142) * fWidth
        self.Save(bSave)

    def SetCellFontColor(self, sheet, strCell, listColor, bSave=False):
        row, col, value, style = self._copyStyle(sheet, strCell)
        style.font.colour_index = self._getColor(listColor)
        tempSheet = self._getWtSheet(sheet)
        tempSheet.write(row, col, value, style)
        self.Save(bSave)

    def SetRangeFontColor(self, sheet, strRange, listColor, bSave=False):
        min_col, min_row, max_col, max_row = self._getCells(strRange)
        for row in range(min_row, max_row):
            for col in range(min_col, max_col):
                self.SetCellFontColor(sheet, [row + 1, col + 1], listColor, bSave)

    def SetCellColor(self, sheet, strCell, listColor, bSave=False):
        row, col, value, style = self._copyStyle(sheet, strCell)
        style.pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        style.pattern.pattern_fore_colour = self._getColor(listColor)
        tempSheet = self._getWtSheet(sheet)
        tempSheet.write(row, col, value, style)
        self.Save(bSave)

    def SetRangeColor(self, sheet, strRange, listColor, bSave=False):
        min_col, min_row, max_col, max_row = self._getCells(strRange)
        for row in range(min_row, max_row):
            for col in range(min_col, max_col):
                self.SetCellColor(sheet, [row + 1, col + 1], listColor, bSave)

    def _getWtSheet(self, sheet):
        return self._Excel03__wt.get_sheet(sheet)

    def _getRdSheet(self, sheet, isrefresh=True):
        if isrefresh:
            self.CloseExcel(True)
            self.OpenExcel(self._Excel03__path)
        if type(sheet) == str:
            return self._Excel03__rd.sheet_by_name(sheet)
        if type(sheet) == int:
            return self._Excel03__rd.sheet_by_index(sheet)

    def _getCells(self, strCell):
        if type(strCell) == str:
            expTemp = strCell.upper()
            if ':' in expTemp:
                min_col, min_row, max_col, max_row = utils.cell.range_boundaries(expTemp)
                return (
                 min_col - 1, min_row - 1, max_col, max_row)
            column, row = utils.cell.coordinate_from_string(expTemp)
            column = utils.column_index_from_string(column)
            return (
             column - 1, row - 1)
        else:
            if type(strCell) == list and len(strCell) == 2:
                valueTemp = strCell[0]
                if type(valueTemp) == int:
                    return (strCell[1] - 1, strCell[0] - 1)
                if type(valueTemp) == list:
                    return (strCell[0][1] - 1, strCell[0][0] - 1, strCell[1][1], strCell[1][0])
                self._getException(strCell)
            else:
                self._getException(strCell)

    def _copyStyle(self, sheet, strCell, isrefresh=True):
        tempSheet = self._getRdSheet(sheet, isrefresh)
        col, row = self._getCells(strCell)
        max_col = tempSheet.ncols
        max_row = tempSheet.nrows
        if col < max_col:
            if row < max_row:
                style = self._Excel03__style_list[tempSheet.cell_xf_index(row, col)]
                value = tempSheet.cell_value(row, col)
                return (
                 row, col, value, style)
        return (row, col, '', xlwt.Style.default_style)

    def _getColor(self, color):
        colour_mark = None
        if type(color) == str and len(color) == 6:
            colour_mark = 'COR_' + color.upper()
            color = [int('0x' + color[i:i + 2].upper(), 16) for i in range(0, len(color), 2)]
        else:
            if type(color) == list:
                if len(color) == 3:
                    tempList = list(filter(lambda x: type(x) == int and 0 <= x <= 255, color))
                    if len(tempList) == 3:
                        colour_mark = 'COR_' + ''.join([str(hex(r))[-2:].replace('x', '0').upper() for r in color])
                else:
                    self._getException(color)
            else:
                self._getException(color)
        colour_mark = colour_mark.lower()
        if colour_mark:
            if colour_mark not in xlwt.Style.colour_map.keys():
                for i in range(8, 63):
                    if i not in xlwt.Style.colour_map.values():
                        xlwt.add_palette_colour(colour_mark, i)
                        self._Excel03__wt.set_colour_RGB(i, color[0], color[1], color[2])
                        break

            return xlwt.Style.colour_map[colour_mark] - 8
        self._getException(color)

    def __handleSheet(self, sheet, wtSheetname=None, mvIndex=-1, option=2):
        self.CloseExcel(True)
        self._Excel03__rd = xlrd.open_workbook((self._Excel03__path), formatting_info=True)
        w = HandleSheetFilter(rdSheetName=sheet, wtSheetname=wtSheetname, mvIndex=mvIndex, option=option)
        process(XLRDReader(self._Excel03__rd, 'unknown.xls'), w)
        self._Excel03__wt = w.output[0][1]
        self._Excel03__style_list = w.style_list

    def __handleCell(self, sheet, cellRange, way=True, option=True):
        self.CloseExcel(True)
        self._Excel03__rd = xlrd.open_workbook((self._Excel03__path), formatting_info=True)
        w = HandleCellFilter(rdSheetName=sheet, cellRange=cellRange, way=way, option=option)
        process(XLRDReader(self._Excel03__rd, 'unknown.xls'), w)
        self._Excel03__wt = w.output[0][1]
        self._Excel03__style_list = w.style_list

    def __handleMerge(self, sheet, cellRange):
        self.CloseExcel(True)
        self._Excel03__rd = xlrd.open_workbook((self._Excel03__path), formatting_info=True)
        w = HandleMergeFilter(rdSheetName=sheet, cellRange=cellRange)
        process(XLRDReader(self._Excel03__rd, 'unknown.xls'), w)
        self._Excel03__wt = w.output[0][1]
        self._Excel03__style_list = w.style_list

    def _getException(self, msg):
        raise Exception('参数错误：%s' % str(msg))


class HandleSheetFilter(BaseFilter, BaseWriter):
    __doc__ = '\n    sheet表操作\n    rdSheetName：要操作的表名\n    wtSheetname：复制后的表名\n    option：1：移动；2：复制，3：删除\n    '

    def __init__(self, rdSheetName, wtSheetname=None, mvIndex=-1, option=2):
        self._HandleSheetFilter__rdSheetName = rdSheetName
        self._HandleSheetFilter__wtSheetname = wtSheetname
        self._HandleSheetFilter__pading_sheet = None
        self._HandleSheetFilter__option = option
        self._HandleSheetFilter__mvIndex = mvIndex
        self._HandleSheetFilter__sheetIndex = -1
        self.output = []

    def sheetIndex(self, name):
        sheetNames = self.rdbook.sheet_names()
        for i in range(len(sheetNames)):
            if sheetNames[i] == name:
                return i

        return -1

    def workbook(self, rdbook, wtbook_name):
        self.rdbook = rdbook
        if type(self._HandleSheetFilter__rdSheetName) == int:
            self._HandleSheetFilter__sheetIndex = self._HandleSheetFilter__rdSheetName
            self._HandleSheetFilter__pading_sheet = self.rdbook.sheet_by_index(self._HandleSheetFilter__rdSheetName)
            self._HandleSheetFilter__rdSheetName = self._HandleSheetFilter__pading_sheet.name
        else:
            self._HandleSheetFilter__pading_sheet = self.rdbook.sheet_by_name(self._HandleSheetFilter__rdSheetName)
            self._HandleSheetFilter__sheetIndex = self.sheetIndex(self._HandleSheetFilter__rdSheetName)
        BaseWriter.workbook(self, rdbook, wtbook_name)

    def sheet(self, rdsheet, wtsheet_name):
        self.rdsheet = rdsheet
        self.wtsheet_name = wtsheet_name
        if self._HandleSheetFilter__sheetIndex > 0:
            if self._HandleSheetFilter__mvIndex > 0 and self._HandleSheetFilter__sheetIndex != self._HandleSheetFilter__mvIndex:
                myIndex = self.sheetIndex(rdsheet.name)
                if myIndex == self._HandleSheetFilter__mvIndex:
                    if self._HandleSheetFilter__pading_sheet is None:
                        self._HandleSheetFilter__pading_sheet = self.rdbook.sheet_by_name(self._HandleSheetFilter__rdSheetName)
                    BaseWriter.sheet(self, self._HandleSheetFilter__pading_sheet, self._HandleSheetFilter__rdSheetName)
                    for row_x in xrange(self._HandleSheetFilter__pading_sheet.nrows):
                        BaseWriter.row(self, row_x, row_x)
                        for col_x in xrange(self._HandleSheetFilter__pading_sheet.row_len(row_x)):
                            BaseWriter.cell(self, row_x, col_x, row_x, col_x)

                    BaseWriter.sheet(self, rdsheet, wtsheet_name)
            else:
                if self._HandleSheetFilter__rdSheetName == rdsheet.name:
                    self._HandleSheetFilter__pading_sheet = rdsheet
                    if self._HandleSheetFilter__option == 2:
                        BaseWriter.sheet(self, self._HandleSheetFilter__pading_sheet, self._HandleSheetFilter__wtSheetname)
                else:
                    BaseWriter.sheet(self, rdsheet, wtsheet_name)
        else:
            if self._HandleSheetFilter__rdSheetName == rdsheet.name:
                self._HandleSheetFilter__pading_sheet = rdsheet
            if self._HandleSheetFilter__option < 3 or self._HandleSheetFilter__rdSheetName != rdsheet.name:
                BaseWriter.sheet(self, rdsheet, wtsheet_name)
        if self._HandleSheetFilter__option == 2:
            if rdsheet.name == self.rdbook.sheet_by_index(-1).name:
                BaseWriter.sheet(self, self._HandleSheetFilter__pading_sheet, self._HandleSheetFilter__wtSheetname)
                for row_x in xrange(self._HandleSheetFilter__pading_sheet.nrows):
                    BaseWriter.row(self, row_x, row_x)
                    for col_x in xrange(self._HandleSheetFilter__pading_sheet.row_len(row_x)):
                        BaseWriter.cell(self, row_x, col_x, row_x, col_x)

    def close(self):
        if self.wtbook is not None:
            self.output.append((self.wtname, self.wtbook))
            del self.wtbook


class HandleCellFilter(BaseFilter, BaseWriter):
    __doc__ = '\n    cell 的位移操作\n    rdSheetName：sheet表名\n    cellRange：操作区域\n    way: True:横向，False：纵向\n    option：True：增，False：删\n    '

    def __init__(self, rdSheetName, cellRange, way=True, option=True):
        self._HandleCellFilter__rdSheetName = rdSheetName
        self._HandleCellFilter__n_cells = cellRange
        self._HandleCellFilter__n_row = cellRange[2] - cellRange[0]
        self._HandleCellFilter__n_col = cellRange[3] - cellRange[1]
        self._HandleCellFilter__way = way
        self._HandleCellFilter__option = option
        self.output = []

    def workbook(self, rdbook, wtbook_name):
        self.rdbook = rdbook
        if type(self._HandleCellFilter__rdSheetName) == int:
            self._HandleCellFilter__rdSheetName = self.rdbook.sheet_by_index(self._HandleCellFilter__rdSheetName).name
        BaseWriter.workbook(self, rdbook, wtbook_name)

    def sheet(self, rdsheet, wtsheet_name):
        self.rdsheet = rdsheet
        BaseWriter.sheet(self, rdsheet, wtsheet_name)

    def cell(self, rdrowx, rdcolx, wtrowx, wtcolx):
        if self.rdsheet.name == self._HandleCellFilter__rdSheetName:
            if self._HandleCellFilter__option:
                if self._HandleCellFilter__way:
                    if rdcolx >= self._HandleCellFilter__n_cells[1]:
                        if self._HandleCellFilter__n_cells[0] <= rdrowx < self._HandleCellFilter__n_cells[2]:
                            BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx + self._HandleCellFilter__n_col)
                else:
                    BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx)
            elif rdrowx >= self._HandleCellFilter__n_cells[0]:
                if self._HandleCellFilter__n_cells[1] <= rdcolx < self._HandleCellFilter__n_cells[3]:
                    BaseWriter.cell(self, rdrowx, rdcolx, wtrowx + self._HandleCellFilter__n_row, wtcolx)
                else:
                    BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx)
            elif self._HandleCellFilter__way:
                if rdcolx >= self._HandleCellFilter__n_cells[3]:
                    if self._HandleCellFilter__n_cells[0] <= rdrowx < self._HandleCellFilter__n_cells[2]:
                        BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx - self._HandleCellFilter__n_col)
                if self._HandleCellFilter__n_cells[1] <= rdcolx < self._HandleCellFilter__n_cells[3]:
                    if self._HandleCellFilter__n_cells[0] <= rdrowx < self._HandleCellFilter__n_cells[2]:
                        pass
                    else:
                        BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx)
        else:
            if rdrowx >= self._HandleCellFilter__n_cells[2]:
                if self._HandleCellFilter__n_cells[1] <= rdcolx < self._HandleCellFilter__n_cells[3]:
                    BaseWriter.cell(self, rdrowx, rdcolx, wtrowx - self._HandleCellFilter__n_row, wtcolx)
                else:
                    if self._HandleCellFilter__n_cells[1] <= rdcolx < self._HandleCellFilter__n_cells[3]:
                        if self._HandleCellFilter__n_cells[0] <= rdrowx < self._HandleCellFilter__n_cells[2]:
                            pass
                        else:
                            BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx)
            else:
                BaseWriter.cell(self, rdrowx, rdcolx, wtrowx, wtcolx)

    def close(self):
        if self.wtbook is not None:
            self.output.append((self.wtname, self.wtbook))
            del self.wtbook


class HandleMergeFilter(BaseFilter, BaseWriter):
    __doc__ = '\n    sheet 拆分单元格\n    rdSheetName：要操作的表名\n    cellRange：要拆分的单元格\n    '

    def __init__(self, rdSheetName, cellRange):
        self._HandleMergeFilter__rdSheetName = rdSheetName
        self._HandleMergeFilter__cellRange = cellRange
        self.output = []

    def workbook(self, rdbook, wtbook_name):
        self.rdbook = rdbook
        if type(self._HandleMergeFilter__rdSheetName) == int:
            self._HandleMergeFilter__rdSheetName = self.rdbook.sheet_by_index(self._HandleMergeFilter__rdSheetName).name
        BaseWriter.workbook(self, rdbook, wtbook_name)

    def sheet(self, rdsheet, wtsheet_name):
        self.rdsheet = rdsheet
        BaseWriter.sheet(self, rdsheet, wtsheet_name)
        if rdsheet.name == self._HandleMergeFilter__rdSheetName:
            del self.merged_cell_top_left_map[(self._HandleMergeFilter__cellRange[1], self._HandleMergeFilter__cellRange[0])]

    def close(self):
        if self.wtbook is not None:
            self.output.append((self.wtname, self.wtbook))
            del self.wtbook


# if __name__ == '__main__':
    # pass
# okay decompiling Excel03.pyc
